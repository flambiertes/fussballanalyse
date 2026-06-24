"""
WM-2026-Simulation v3
Architektur: N vollstaendige Turnier-Simulationen
  1. Gruppenphase → eine feste Tabelle pro Gruppe
  2. Beste 8 Dritte bestimmen → feste KO-Bracket-Zuteilung
  3. KO-Runden → jedes Team genau einmal pro Simulation
  4. 1000x wiederholen → Haeufigkeiten zaehlen

Ausgabe: Excel mit
  - Uebersicht (alle 48 Teams)
  - Gruppe A-L (Tabelle + Spiele)
  - Turnierbaum (Sechzehntelfinale ... Finale)
  - Drittplatzierte
"""

import csv
import json
from datetime import datetime
import numpy as np
import pandas as pd
from scipy.stats import poisson
from pathlib import Path
from collections import defaultdict
from tqdm.auto import tqdm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent / "data"
N_TOURNAMENTS = 5000
MARKET_ATTACK_LOG_ADJ = 0.35
MARKET_DEFENSE_LOG_ADJ = 0.35
ELO_MAX_GOAL_BOOST = 0.08
SCORE_TIP_MODE = "average"  # "expected_points", "most_common" oder "average"
TIP_SCORE_MAX_GOALS = 5


def _load_model_metadata() -> dict:
    """Laedt mu und home_adv aus model_metadata.json (Fallback: 0.0 wenn nicht vorhanden)."""
    path = DATA_DIR / "model_metadata.json"
    if path.exists():
        with open(path) as f:
            return json.load(f)
    return {"mu": 0.0, "home_adv": 0.25}


_MODEL_META = _load_model_metadata()
MU_INTERCEPT = _MODEL_META.get("mu", 0.0) + 0.302  # Globaler Intercept aus Poisson-Fitting

# ---------------------------------------------------------------------------
# Fester KO-Spielplan: Slot-Beschreibung -> Bracket-Position
# ---------------------------------------------------------------------------
# Annex C der offiziellen FIFA-Regularien ordnet jede Kombination der acht
# qualifizierten Gruppendritten eindeutig einem Gruppensieger zu.
THIRD_PLACE_COMBINATIONS = DATA_DIR / "third_place_combinations.csv"
THIRD_SLOTS_BY_WINNER = {
    "1A": "3rd Group C/E/F/H/I",
    "1B": "3rd Group E/F/G/I/J",
    "1D": "3rd Group B/E/F/I/J",
    "1E": "3rd Group A/B/C/D/F",
    "1G": "3rd Group A/E/H/I/J",
    "1I": "3rd Group C/D/F/G/H",
    "1K": "3rd Group D/E/I/J/L",
    "1L": "3rd Group E/H/I/J/K",
}

# Turnierbaum-Layout: match_nr -> (row_team1, row_team2, col_team, col_score)
BRACKET = {
    73: (2,  3,  1, 2),  76: (34, 35, 1, 2),
    75: (6,  7,  1, 2),  78: (38, 39, 1, 2),
    74: (10, 11, 1, 2),  79: (42, 43, 1, 2),
    77: (14, 15, 1, 2),  80: (46, 47, 1, 2),
    83: (18, 19, 1, 2),  86: (50, 51, 1, 2),
    84: (22, 23, 1, 2),  88: (54, 55, 1, 2),
    81: (26, 27, 1, 2),  85: (58, 59, 1, 2),
    82: (30, 31, 1, 2),  87: (62, 63, 1, 2),
    90: (4,  5,  4, 5),  91: (36, 37, 4, 5),
    89: (12, 13, 4, 5),  92: (44, 45, 4, 5),
    93: (20, 21, 4, 5),  95: (52, 53, 4, 5),
    94: (28, 29, 4, 5),  96: (60, 61, 4, 5),
    97:  (8,  9,  7, 8),  99:  (40, 41, 7, 8),
    98:  (24, 25, 7, 8),  100: (56, 57, 7, 8),
    101: (16, 17, 10, 11), 102: (48, 49, 10, 11),
    104: (32, 33, 13, 14),
    103: (37, 38, 13, 14),
}

STAGE_OF_MATCH = {mnr: "p_R16" for mnr in range(73, 89)}
STAGE_OF_MATCH.update({mnr: "p_QF" for mnr in range(89, 97)})
STAGE_OF_MATCH.update({mnr: "p_SF" for mnr in range(97, 101)})
STAGE_OF_MATCH.update({101: "p_F", 102: "p_F", 104: "p_Winner"})


# ---------------------------------------------------------------------------
# Datenladen
# ---------------------------------------------------------------------------
def load_inputs():
    strengths = pd.read_csv(DATA_DIR / "team_strengths.csv").set_index("team")
    groups_df = pd.read_csv(DATA_DIR / "wm2026_groups.csv")
    groups    = {g: list(v["team"]) for g, v in groups_df.groupby("group")}
    ko_df     = pd.read_csv(DATA_DIR / "wm2026_matches_knockout.csv").sort_values("match_nr")
    return strengths, groups, ko_df


def load_fixed_group_matches():
    """Bereits gespielte Gruppenspiele als {grp: {(home,away): (gh,ga)}}."""
    path = DATA_DIR / "wm2026_matches_group.csv"
    result = defaultdict(dict)
    if not path.exists(): return result
    df = pd.read_csv(path, parse_dates=["date"])
    if "goals_home" not in df.columns: return result
    played = df.dropna(subset=["goals_home","goals_away"])
    for _, row in played.iterrows():
        result[row["group"]][(row["team_home"], row["team_away"])] = (int(row["goals_home"]), int(row["goals_away"]))
    return result


def load_fixed_ko_matches():
    """Bereits gespielte KO-Spiele als {match_nr: (team_a, team_b, ga, gb)}."""
    path = DATA_DIR / "wm2026_matches_knockout.csv"
    result = {}
    if not path.exists(): return result
    df = pd.read_csv(path)
    if "goals_home" not in df.columns: return result
    played = df.dropna(subset=["goals_home","goals_away"])
    for _, row in played.iterrows():
        mnr = int(row["match_nr"])
        ga, gb = int(row["goals_home"]), int(row["goals_away"])
        winner = row["team_home"] if ga > gb else row["team_away"]
        result[mnr] = {"team_a": row["team_home"], "team_b": row["team_away"],
                       "goals_a": ga, "goals_b": gb, "winner": winner, "extra": ""}
    return result


# ---------------------------------------------------------------------------
# Spielsimulation
# ---------------------------------------------------------------------------
def expected_goals(team_a, team_b, strengths, neutral=True):
    def get(t, col, d):
        return strengths.at[t, col] if (t in strengths.index and col in strengths.columns) else d
    if "att" in strengths.columns:
        h_adv = 0.0 if neutral else get(team_a, "home_adv", 0.25)
        mv_att_a = get(team_a, "mv_att_norm", get(team_a, "mv_norm", 0.5))
        mv_att_b = get(team_b, "mv_att_norm", get(team_b, "mv_norm", 0.5))
        mv_def_a = get(team_a, "mv_def_norm", get(team_a, "mv_norm", 0.5))
        mv_def_b = get(team_b, "mv_def_norm", get(team_b, "mv_norm", 0.5))
        elo_a = get(team_a, "elo_norm", 0.5)
        elo_b = get(team_b, "elo_norm", 0.5)
        elo_boost_a = np.log(1 + ELO_MAX_GOAL_BOOST * max(elo_a - elo_b, 0))
        elo_boost_b = np.log(1 + ELO_MAX_GOAL_BOOST * max(elo_b - elo_a, 0))

        log_lam_a = (
            get(team_a, "att", 0) +
            get(team_b, "def", 0) +
            h_adv +
            MARKET_ATTACK_LOG_ADJ * (mv_att_a - 0.5) -
            MARKET_DEFENSE_LOG_ADJ * (mv_def_b - 0.5) +
            elo_boost_a +
            MU_INTERCEPT
        )
        log_lam_b = (
            get(team_b, "att", 0) +
            get(team_a, "def", 0) +
            MARKET_ATTACK_LOG_ADJ * (mv_att_b - 0.5) -
            MARKET_DEFENSE_LOG_ADJ * (mv_def_a - 0.5) +
            elo_boost_b +
            MU_INTERCEPT
        )
        lam_a = np.exp(np.clip(log_lam_a, -6, 6))
        lam_b = np.exp(np.clip(log_lam_b, -6, 6))
    else:
        s_a = get(team_a,"strength_combined",0.5); s_b = get(team_b,"strength_combined",0.5)
        lam_a = 2.75 * s_a / (s_a + s_b + 1e-9)
        lam_b = 2.75 - lam_a
    return max(lam_a, 0.05), max(lam_b, 0.05)


def sim_match(lam_a, lam_b):
    return int(poisson.rvs(lam_a)), int(poisson.rvs(lam_b))


def sim_ko_match(team_a, team_b, strengths):
    """90 Min + Verlaengerung + Elfmeter. Gibt (winner, ga, gb, extra) zurueck."""
    lam_a, lam_b = expected_goals(team_a, team_b, strengths)
    g_a, g_b = sim_match(lam_a, lam_b)
    if g_a != g_b:
        return (team_a if g_a > g_b else team_b), g_a, g_b, ""
    et_a, et_b = sim_match(lam_a / 3, lam_b / 3)
    g_a += et_a; g_b += et_b
    if g_a != g_b:
        return (team_a if g_a > g_b else team_b), g_a, g_b, "n.V."
    def get(t, col, d):
        return strengths.at[t, col] if t in strengths.index and col in strengths.columns else d
    s_a = get(team_a,"strength_combined",0.5); s_b = get(team_b,"strength_combined",0.5)
    p_a = np.clip(s_a / (s_a + s_b + 1e-9), 0.40, 0.60)
    return (team_a if np.random.random() < p_a else team_b), g_a, g_b, "n.E."


# ---------------------------------------------------------------------------
# Gruppenphase (eine Simulation)
# ---------------------------------------------------------------------------
def simulate_group_once(teams, strengths, fixed):
    """Eine Gruppenphase. Gibt (ranking, match_scores, records) zurueck."""
    records = {t: {"pts":0,"gf":0,"ga":0} for t in teams}
    match_scores = {}
    for i, h in enumerate(teams):
        for a in teams[i+1:]:
            if (h, a) in fixed:
                gh, ga = fixed[(h, a)]
            elif (a, h) in fixed:
                ga, gh = fixed[(a, h)]
            else:
                lam_h, lam_a = expected_goals(h, a, strengths)
                gh, ga = sim_match(lam_h, lam_a)
            match_scores[(h, a)] = (gh, ga)
            records[h]["gf"] += gh; records[h]["ga"] += ga
            records[a]["gf"] += ga; records[a]["ga"] += gh
            if gh > ga:    records[h]["pts"] += 3
            elif gh == ga: records[h]["pts"] += 1; records[a]["pts"] += 1
            else:          records[a]["pts"] += 3
    table = pd.DataFrame.from_dict(records, orient="index")
    table["gd"] = table["gf"] - table["ga"]
    table = table.sort_values(["pts","gd","gf"], ascending=False)
    return list(table.index), match_scores, records


# ---------------------------------------------------------------------------
# Bracket-Zuteilung (deterministisch aus Gruppenerg.)
# ---------------------------------------------------------------------------
def assign_thirds_to_slots(best8_thirds):
    """
    Weist die 8 besten Dritten gemaess FIFA-Regularien Annex C den Slots zu.
    best8_thirds: [(team, pts, gd, gf, grp), ...] sortiert nach Staerke.
    Gibt {slot_description: team} zurueck.
    """
    combinations = pd.read_csv(THIRD_PLACE_COMBINATIONS, dtype=str)
    teams_by_group = {grp: team for team, _, _, _, grp in best8_thirds}
    qualified_groups = "".join(sorted(teams_by_group))
    matches = combinations[combinations["qualified_groups"] == qualified_groups]
    if len(matches) != 1:
        raise ValueError(f"Keine eindeutige FIFA-Zuteilung fuer Dritte aus {qualified_groups}")
    row = matches.iloc[0]
    return {
        slot: teams_by_group[row[f"third_for_{winner}"]]
        for winner, slot in THIRD_SLOTS_BY_WINNER.items()
    }


def build_qualifiers(group_standings, all_records):
    """Baut {slot_desc: team} aus den Gruppenergebnissen."""
    qualifiers = {}
    thirds_info = []

    for grp, (ranking, records) in group_standings.items():
        qualifiers[f"Winner Group {grp}"]    = ranking[0]
        qualifiers[f"Runner-up Group {grp}"] = ranking[1]
        if len(ranking) >= 3:
            t = ranking[2]
            thirds_info.append((t, records[t]["pts"],
                                 records[t]["gf"] - records[t]["ga"],
                                 records[t]["gf"], grp))

    thirds_info.sort(key=lambda x: (-x[1], -x[2], -x[3]))
    best8 = thirds_info[:8]
    third_assignments = assign_thirds_to_slots(best8)
    qualifiers.update(third_assignments)
    return qualifiers


# ---------------------------------------------------------------------------
# Eine vollstaendige Turnier-Simulation
# ---------------------------------------------------------------------------
def resolve_team(desc, qualifiers, ko_results, ko_losers=None):
    if desc.startswith("Winner Match "):
        prev = int(desc.split()[-1])
        return ko_results.get(prev, desc)
    if desc.startswith("Loser Match "):
        prev = int(desc.split()[-1])
        return (ko_losers or {}).get(prev, desc)
    return qualifiers.get(desc, desc)


def simulate_one_tournament(groups, strengths, fixed_group_matches, ko_df, fixed_ko):
    """
    Eine vollstaendige WM-Simulation. Jedes Team erscheint genau einmal.
    Gibt (group_standings, qualifiers, ko_results, ko_scores) zurueck.
    """
    # Gruppenphase
    group_standings = {}
    all_match_scores = {}
    for grp, teams in groups.items():
        fixed = fixed_group_matches.get(grp, {})
        ranking, scores, records = simulate_group_once(teams, strengths, fixed)
        group_standings[grp] = (ranking, records)
        all_match_scores.update(scores)

    # Bracket-Zuteilung
    qualifiers = build_qualifiers(group_standings, all_match_scores)

    # KO-Phase
    ko_results = {}  # match_nr -> winner
    ko_losers  = {}  # match_nr -> loser
    ko_scores  = {}  # match_nr -> (team_a, team_b, ga, gb, extra)

    for _, row in ko_df.iterrows():
        mnr  = int(row["match_nr"])
        if mnr in fixed_ko:
            fk = fixed_ko[mnr]
            ko_results[mnr] = fk["winner"]
            ko_losers[mnr]   = fk["team_b"] if fk["winner"] == fk["team_a"] else fk["team_a"]
            ko_scores[mnr]  = (fk["team_a"], fk["team_b"], fk["goals_a"], fk["goals_b"], fk["extra"])
            continue
        ta = resolve_team(row["team_home_desc"], qualifiers, ko_results, ko_losers)
        tb = resolve_team(row["team_away_desc"], qualifiers, ko_results, ko_losers)
        if "Group" in ta or "Match" in ta or "Group" in tb or "Match" in tb:
            continue  # Abhaengigkeit noch nicht aufgeloest
        winner, ga, gb, extra = sim_ko_match(ta, tb, strengths)
        ko_results[mnr] = winner
        ko_losers[mnr]  = tb if winner == ta else ta
        ko_scores[mnr]  = (ta, tb, ga, gb, extra)

    return group_standings, all_match_scores, qualifiers, ko_results, ko_scores


# ---------------------------------------------------------------------------
# N Turnier-Simulationen
# ---------------------------------------------------------------------------
def run_n_tournaments(groups, strengths, fixed_group_matches, ko_df, fixed_ko, n=N_TOURNAMENTS):
    """N vollstaendige Turnier-Simulationen. Zaehlt Haeufigkeiten."""
    group_pos_counts = {grp: {t: [0,0,0,0] for t in teams} for grp, teams in groups.items()}
    team_stage_counts = defaultdict(lambda: defaultdict(int))
    ko_match_stats = defaultdict(lambda: {
        "team_a_counts": defaultdict(int), "team_b_counts": defaultdict(int),
        "winner_counts": defaultdict(int),
        "goals_a": [], "goals_b": [],
        "score_counts": defaultdict(int),
        "score_counts_win_a": defaultdict(int),
        "score_counts_win_b": defaultdict(int),
    })
    # Durchschnittliche Gruppenspiel-Scores
    group_match_scores = defaultdict(lambda: {"goals_h": [], "goals_a": [],
                                               "wins_h": 0, "draws": 0, "wins_a": 0,
                                               "score_counts": defaultdict(int)})
    tournament_paths = []

    for tournament_id in tqdm(range(1, n + 1), desc=f"Simuliere {n} Turniere"):
        g_standings, g_scores, qualifiers, ko_results, ko_scores = simulate_one_tournament(
            groups, strengths, fixed_group_matches, ko_df, fixed_ko
        )
        tournament_paths.append({
            "tournament_id": tournament_id,
            "group_standings": g_standings,
            "group_scores": g_scores,
            "qualifiers": qualifiers,
            "ko_results": ko_results,
            "ko_scores": ko_scores,
        })

        # Gruppenphase zaehlen
        for grp, (ranking, _) in g_standings.items():
            for pos, team in enumerate(ranking[:4]):
                group_pos_counts[grp][team][pos] += 1

        # Gruppenspiel-Scores
        for grp, teams in groups.items():
            fixed = fixed_group_matches.get(grp, {})
            _, match_scores, _ = simulate_group_once(teams, strengths, fixed)
            for (h, a), (gh, ga) in match_scores.items():
                group_match_scores[(h, a)]["goals_h"].append(gh)
                group_match_scores[(h, a)]["goals_a"].append(ga)
                group_match_scores[(h, a)]["score_counts"][(gh, ga)] += 1
                if gh > ga:    group_match_scores[(h, a)]["wins_h"] += 1
                elif gh == ga: group_match_scores[(h, a)]["draws"]  += 1
                else:          group_match_scores[(h, a)]["wins_a"] += 1

        # R32-Qualifikation tracken (wer ist im Bracket?)
        for slot_desc, team in qualifiers.items():
            team_stage_counts[team]["p_R32_entry"] += 1
            if "3rd" in slot_desc:
                team_stage_counts[team]["p_R32_as_third"] += 1

        # KO-Phase zaehlen
        for mnr, winner in ko_results.items():
            stage = STAGE_OF_MATCH.get(mnr)
            if stage:
                team_stage_counts[winner][stage] += 1

        for mnr, (ta, tb, ga, gb, extra) in ko_scores.items():
            s = ko_match_stats[mnr]
            s["team_a_counts"][ta] += 1
            s["team_b_counts"][tb] += 1
            winner = ko_results.get(mnr)
            if winner:
                s["winner_counts"][winner] += 1
                s["score_counts"][(ga, gb)] += 1
                if winner == ta: s["score_counts_win_a"][(ga, gb)] += 1
                else:            s["score_counts_win_b"][(ga, gb)] += 1
            s["goals_a"].append(ga)
            s["goals_b"].append(gb)

    return group_pos_counts, group_match_scores, ko_match_stats, team_stage_counts, tournament_paths, n


def rank_representative_tournaments(tournament_paths, group_pos_counts, team_stage_counts, n):
    """
    Waehlt den real simulierten Turnierverlauf, der den aggregierten
    Wahrscheinlichkeiten am besten entspricht.
    """
    def score(path):
        value = 0.0
        position_scores = [0.0, 0.0, 0.0, 0.0]
        for grp, (ranking, _) in path["group_standings"].items():
            for pos, team in enumerate(ranking[:4]):
                position_scores[pos] += group_pos_counts[grp][team][pos] / n
        value += sum(pos_score / max(len(path["group_standings"]), 1) for pos_score in position_scores)

        qualified_teams = set(path["qualifiers"].values())
        value += sum(team_stage_counts[team]["p_R32_entry"] / n for team in qualified_teams) / max(len(qualified_teams), 1)

        stage_slots = {
            "p_R16": 16,
            "p_QF": 8,
            "p_SF": 4,
            "p_F": 2,
            "p_Winner": 1,
        }
        stage_scores = defaultdict(float)
        for mnr, winner in path["ko_results"].items():
            stage = STAGE_OF_MATCH.get(mnr)
            if stage:
                stage_scores[stage] += team_stage_counts[winner][stage] / n
        for stage, slots in stage_slots.items():
            value += stage_scores[stage] / slots
        return value

    ranked = []
    for path in tournament_paths:
        path["matching_score"] = score(path)
        ranked.append(path)
    return sorted(ranked, key=lambda path: -path["matching_score"])


def get_podium(path):
    """Gibt Weltmeister, Zweiten, Dritten und Vierten eines Turnierverlaufs zurueck."""
    champion = path["ko_results"][104]
    final_a, final_b, _, _, _ = path["ko_scores"][104]
    runner_up = final_b if champion == final_a else final_a
    third = path["ko_results"][103]
    third_a, third_b, _, _, _ = path["ko_scores"][103]
    fourth = third_b if third == third_a else third_a
    return champion, runner_up, third, fourth


# ---------------------------------------------------------------------------
# Hilfsfunktionen fuer Excel
# ---------------------------------------------------------------------------
def score_from_avg(avg_h, avg_a, p_win_h, p_win_a):
    """Rundet Ø-Tore zu ganzen Zahlen, konsistent mit wahrscheinlichstem Ausgang."""
    gh, ga = round(avg_h), round(avg_a)
    if p_win_h > 0.55 and gh <= ga: gh = ga + 1
    elif p_win_a > 0.55 and ga <= gh: ga = gh + 1
    return gh, ga


def most_common_score(score_counts):
    """Gibt den haeufigsten simulierten Score zurueck."""
    if not score_counts:
        return None
    return max(
        score_counts.items(),
        key=lambda item: (item[1], -sum(item[0]), -abs(item[0][0] - item[0][1])),
    )[0]


def tip_points(tip, actual):
    """Punkte fuer einen Tipp nach Ergebnis/Differenz/Tendenz."""
    th, ta = tip
    ah, aa = actual
    if tip == actual:
        return 4
    tip_diff = th - ta
    actual_diff = ah - aa
    if tip_diff == actual_diff:
        return 3
    if (tip_diff > 0 and actual_diff > 0) or (tip_diff < 0 and actual_diff < 0) or (tip_diff == 0 and actual_diff == 0):
        return 2
    return 0


def expected_points_score(score_counts, max_goals=TIP_SCORE_MAX_GOALS):
    """Waehlt den Tipp mit maximalem Erwartungswert nach Tipp-Punktelogik."""
    if not score_counts:
        return None
    candidates = [(h, a) for h in range(max_goals + 1) for a in range(max_goals + 1)]
    return max(
        candidates,
        key=lambda tip: (
            sum(tip_points(tip, actual) * count for actual, count in score_counts.items()),
            score_counts.get(tip, 0),
            -sum(tip),
            -abs(tip[0] - tip[1]),
        ),
    )


def select_tip_score(avg_h, avg_a, p_win_h, p_win_a, score_counts):
    """
    Waehlt den angezeigten Tipp-Score.
    SCORE_TIP_MODE='expected_points': hoechster Erwartungswert nach Tipp-Punkten.
    SCORE_TIP_MODE='most_common': haeufigster simulierter Score.
    SCORE_TIP_MODE='average': alte Logik aus Durchschnittstoren + Rundung.
    Zweiter Rueckgabewert: True, wenn gewaehlter Score und alte Logik gleich sind.
    """
    avg_score = score_from_avg(avg_h, avg_a, p_win_h, p_win_a)
    common_score = most_common_score(score_counts)
    expected_score = expected_points_score(score_counts)
    if SCORE_TIP_MODE == "expected_points" and expected_score is not None:
        return expected_score, expected_score == avg_score
    if SCORE_TIP_MODE == "most_common" and common_score is not None:
        return common_score, common_score == avg_score
    return avg_score, common_score == avg_score if common_score is not None else False


def build_expected_table(teams, avg_scores, fixed_group_matches_flat):
    """Baut konkrete Tabelle aus gerundeten Ø-Ergebnissen (W/U/N/Pts)."""
    records = {t: {"sp":0,"w":0,"d":0,"l":0,"gf":0,"ga":0,"pts":0} for t in teams}
    for (h, a), s in avg_scores.items():
        if h not in teams or a not in teams: continue
        if (h, a) in fixed_group_matches_flat:
            gh, ga = fixed_group_matches_flat[(h, a)]
        elif (a, h) in fixed_group_matches_flat:
            ga, gh = fixed_group_matches_flat[(a, h)]
        else:
            p_wh = s["wins_h"] / max(s["wins_h"] + s["draws"] + s["wins_a"], 1)
            p_wa = s["wins_a"] / max(s["wins_h"] + s["draws"] + s["wins_a"], 1)
            avg_h = np.mean(s["goals_h"]) if s["goals_h"] else 0
            avg_a = np.mean(s["goals_a"]) if s["goals_a"] else 0
            gh, ga = score_from_avg(avg_h, avg_a, p_wh, p_wa)
        records[h]["sp"] += 1; records[a]["sp"] += 1
        records[h]["gf"] += gh; records[h]["ga"] += ga
        records[a]["gf"] += ga; records[a]["ga"] += gh
        if gh > ga:
            records[h]["w"] += 1; records[h]["pts"] += 3; records[a]["l"] += 1
        elif gh == ga:
            records[h]["d"] += 1; records[h]["pts"] += 1
            records[a]["d"] += 1; records[a]["pts"] += 1
        else:
            records[a]["w"] += 1; records[a]["pts"] += 3; records[h]["l"] += 1
    table = [{"team": t, "gd": r["gf"] - r["ga"], **r} for t, r in records.items()]
    table.sort(key=lambda x: (-x["pts"], -x["gd"], -x["gf"]))
    return table


def get_ko_result_summary(ko_match_stats, n):
    """Fasst KO-Match-Statistiken zusammen."""
    results = {}
    for mnr, s in ko_match_stats.items():
        if not s["team_a_counts"]: continue
        best_a = max(s["team_a_counts"], key=s["team_a_counts"].get)
        best_b = max(s["team_b_counts"], key=s["team_b_counts"].get) if s["team_b_counts"] else "?"
        total  = sum(s["winner_counts"].values())
        p_a    = s["winner_counts"].get(best_a, 0) / total if total else 0.5
        avg_ga = np.mean(s["goals_a"]) if s["goals_a"] else 0
        avg_gb = np.mean(s["goals_b"]) if s["goals_b"] else 0
        (ga_disp, gb_disp), score_agrees = select_tip_score(
            avg_ga, avg_gb, p_a, 1 - p_a, s["score_counts"]
        )
        results[mnr] = {
            "team_a": best_a, "team_b": best_b,
            "p_a_wins": round(p_a, 3),
            "score_a": ga_disp, "score_b": gb_disp,
            "score_agrees_with_avg": score_agrees,
        }
    return results


# ---------------------------------------------------------------------------
# Konsistenter Turnierpfad fuer Bracket-Anzeige (kein Team doppelt)
# ---------------------------------------------------------------------------
def get_most_likely_bracket(groups, group_pos_counts, strengths, ko_df, n):
    """
    Berechnet EINEN konsistenten Turnierpfad.
    Jedes Team erscheint genau einmal - kein doppeltes Canada mehr.
    Favorit gewinnt jedes KO-Spiel deterministisch.
    """
    qualifiers = {}
    thirds_info = []

    for grp, teams in groups.items():
        # Erwarteter Platz (gewichteter Durchschnitt der Platzierungen)
        expected_pos = {}
        for t in teams:
            counts = group_pos_counts[grp][t]
            total = max(sum(counts), 1)
            expected_pos[t] = sum(pos * cnt for pos, cnt in enumerate(counts)) / total
        ranked = sorted(teams, key=lambda t: expected_pos[t])
        qualifiers[f"Winner Group {grp}"]    = ranked[0]
        qualifiers[f"Runner-up Group {grp}"] = ranked[1]
        if len(ranked) >= 3:
            t3 = ranked[2]
            thirds_info.append((t3, expected_pos[t3], grp))

    # Beste 8 Dritte
    thirds_info.sort(key=lambda x: x[1])
    best8 = [(t, 0, 0, 0, grp) for t, _, grp in thirds_info[:8]]
    qualifiers.update(assign_thirds_to_slots(best8))

    # KO deterministisch: Favorit gewinnt immer
    ko_results = {}
    ko_losers  = {}
    ko_scores  = {}
    for _, row in ko_df.sort_values("match_nr").iterrows():
        mnr = int(row["match_nr"])
        ta  = resolve_team(row["team_home_desc"], qualifiers, ko_results, ko_losers)
        tb  = resolve_team(row["team_away_desc"], qualifiers, ko_results, ko_losers)
        if "Group" in ta or "Match" in ta or "Group" in tb or "Match" in tb:
            continue
        lam_a, lam_b = expected_goals(ta, tb, strengths)
        ga, gb = round(lam_a), round(lam_b)
        if ga == gb:
            sa = strengths.at[ta,"strength_combined"] if ta in strengths.index else 0.5
            sb = strengths.at[tb,"strength_combined"] if tb in strengths.index else 0.5
            if sa >= sb: ga = gb + 1
            else:        gb = ga + 1
        winner = ta if ga > gb else tb
        ko_results[mnr] = winner
        ko_losers[mnr]  = tb if winner == ta else ta
        ko_scores[mnr]  = (ta, tb, ga, gb, "")

    return qualifiers, ko_results, ko_scores


# ---------------------------------------------------------------------------
# Excel: Erklaerung
# ---------------------------------------------------------------------------
def write_explanation_sheet(wb, strengths):
    ws = wb.create_sheet("Erklaerung")

    def get(team, col, default=0.0):
        return strengths.at[team, col] if team in strengths.index and col in strengths.columns else default

    teams = [t for t in ["France", "Spain"] if t in strengths.index]
    if len(teams) < 2:
        teams = list(strengths.sort_values("strength_combined", ascending=False).head(2).index)
    team_a, team_b = teams[0], teams[1]

    raw_lam_a = np.exp(np.clip(get(team_a, "att") + get(team_b, "def"), -6, 6))
    raw_lam_b = np.exp(np.clip(get(team_b, "att") + get(team_a, "def"), -6, 6))
    mv_att_a = get(team_a, "mv_att_norm", get(team_a, "mv_norm", 0.5))
    mv_att_b = get(team_b, "mv_att_norm", get(team_b, "mv_norm", 0.5))
    mv_def_a = get(team_a, "mv_def_norm", get(team_a, "mv_norm", 0.5))
    mv_def_b = get(team_b, "mv_def_norm", get(team_b, "mv_norm", 0.5))
    elo_a = get(team_a, "elo_norm", 0.5)
    elo_b = get(team_b, "elo_norm", 0.5)
    market_adj_a = MARKET_ATTACK_LOG_ADJ * (mv_att_a - 0.5) - MARKET_DEFENSE_LOG_ADJ * (mv_def_b - 0.5)
    market_adj_b = MARKET_ATTACK_LOG_ADJ * (mv_att_b - 0.5) - MARKET_DEFENSE_LOG_ADJ * (mv_def_a - 0.5)
    elo_boost_a = np.log(1 + ELO_MAX_GOAL_BOOST * max(elo_a - elo_b, 0))
    elo_boost_b = np.log(1 + ELO_MAX_GOAL_BOOST * max(elo_b - elo_a, 0))
    lam_a, lam_b = expected_goals(team_a, team_b, strengths)

    ws.cell(1, 1, "WM-2026-Simulation - Rechenlogik").font = Font(bold=True, size=14)
    ws.cell(2, 1, "Dieses Blatt erklaert, wie aus historischen Torwerten, Elo und Transfermarkt-Werten ein simuliertes Match entsteht.").font = FONT_ITALIC

    sections = [
        ("1. Datenbasis",
         "Das Poisson-Modell wird aus Laenderspielen ab 2015 geschaetzt. Elo wird kumulativ aus historischen Laenderspielen berechnet. Transfermarkt liefert Kaderwerte; wenn Positionswerte vorhanden sind, werden Angriff und Defensive getrennt bewertet."),
        ("2. Poisson-Grundidee",
         "Fuer jedes Team wird eine erwartete Torzahl lambda berechnet. Danach wird die echte Torzahl zufaellig aus einer Poisson-Verteilung gezogen: P(Tore=k) = exp(-lambda) * lambda^k / k!."),
        ("3. Roh-Lambda",
         "Rohwert Team A = exp(Angriff A + Abwehr B). Rohwert Team B = exp(Angriff B + Abwehr A). Der Abwehrparameter ist dabei der Beitrag zur gegnerischen Torerwartung: niedriger ist besser."),
        ("4. Kombinierte Staerke",
         "Die Gesamt-Staerke ist nur noch ein Ranking- und Elfmeterschießen-Anker: 35% Angriff, 25% Defensive, 30% Marktwert, 10% Elo."),
        ("5. Positions-Marktwerte",
         "MW Angriff = 2/3 Mittelfeld + Sturm. MW Abwehr = Torwart + Abwehr + 1/3 Mittelfeld. Damit zaehlt das Mittelfeld anteilig in beide Richtungen."),
        ("6. Anpassung im Match",
         f"Der Marktwert wird direkt im Lambda verrechnet: eigener MW Angriff erhoeht die Torerwartung, gegnerischer MW Abwehr senkt sie. Elo gibt dem Elo-staerkeren Team maximal {int(ELO_MAX_GOAL_BOOST * 100)}% Bonus auf die Torerwartung."),
        ("7. KO-Spiele",
         "In KO-Spielen wird erst das 90-Minuten-Ergebnis gezogen. Bei Gleichstand folgt eine simulierte Verlaengerung mit lambda/3. Ist es danach noch gleich, entscheidet ein Elfmeterspiel mit leichtem Vorteil fuer das staerkere Team."),
        ("8. Turnierauswahl",
         "Alle Turniere werden komplett simuliert. Danach wird ein Referenzturnier gewaehlt: Der Matching-Score bewertet, wie gut ein konkreter Verlauf zu den Gesamtwahrscheinlichkeiten aller Simulationen passt."),
    ]

    row = 4
    for title, text in sections:
        ws.cell(row, 1, title).font = FONT_BOLD
        ws.cell(row, 2, text).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 2
    ws.cell(row, 1, f"Beispielmatch: {team_a} vs {team_b}").font = Font(bold=True, size=12)
    row += 1
    headers = ["Kennzahl", team_a, team_b, "Kommentar"]
    for ci, header in enumerate(headers, 1):
        hdr(ws, row, ci, header)
    row += 1

    example_rows = [
        ("Angriff", get(team_a, "att"), get(team_b, "att"), "Poisson-Angriffsparameter aus historischen Spielen"),
        ("Abwehr", get(team_a, "def"), get(team_b, "def"), "Beitrag zur gegnerischen Torerwartung; niedriger ist besser"),
        ("Ang.-Norm", get(team_a, "att_norm"), get(team_b, "att_norm"), "Angriff auf 0 bis 1 normiert"),
        ("Def.-Norm", get(team_a, "def_quality_norm", 0), get(team_b, "def_quality_norm", 0), "Defensive auf 0 bis 1 normiert; hoeher ist besser"),
        ("Elo", get(team_a, "elo"), get(team_b, "elo"), "Historisches Elo-Rating"),
        ("Elo-Norm", get(team_a, "elo_norm"), get(team_b, "elo_norm"), "Elo auf 0 bis 1 normiert"),
        ("Marktwert-Norm", get(team_a, "mv_norm", 0.5), get(team_b, "mv_norm", 0.5), "Gesamtmarktwert auf 0 bis 1 normiert"),
        ("MW Angriff Norm", get(team_a, "mv_att_norm", get(team_a, "mv_norm", 0.5)), get(team_b, "mv_att_norm", get(team_b, "mv_norm", 0.5)), "Positionsmarktwert Angriff, falls vorhanden"),
        ("MW Abwehr Norm", get(team_a, "mv_def_norm", get(team_a, "mv_norm", 0.5)), get(team_b, "mv_def_norm", get(team_b, "mv_norm", 0.5)), "Positionsmarktwert Defensive, falls vorhanden"),
        ("Gesamt-Staerke", get(team_a, "strength_combined", 0.5), get(team_b, "strength_combined", 0.5), "Gewichtete Gesamtstaerke"),
        ("Roh-lambda", raw_lam_a, raw_lam_b, "exp(Angriff eigenes Team + Abwehr Gegner)"),
        ("Marktwert-Log-Korrektur", market_adj_a, market_adj_b, "Eigener MW Angriff minus gegnerischer MW Abwehr, jeweils um 0,5 zentriert"),
        ("Elo-Log-Bonus", elo_boost_a, elo_boost_b, "Maximal 8% Lambda-Bonus fuer das Elo-staerkere Team"),
        ("Finales lambda", lam_a, lam_b, "Erwartete Tore nach Anpassung"),
    ]
    for label, va, vb, comment in example_rows:
        ws.cell(row, 1, label).alignment = ALIGN_L
        ws.cell(row, 2, round(float(va), 4)).alignment = ALIGN_C
        ws.cell(row, 3, round(float(vb), 4)).alignment = ALIGN_C
        ws.cell(row, 4, comment).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 2
    ws.cell(row, 1, "Poisson-Torverteilung fuer das Beispiel").font = Font(bold=True, size=12)
    row += 1
    for ci, header in enumerate(["Tore", f"P({team_a})", f"P({team_b})"], 1):
        hdr(ws, row, ci, header)
    row += 1
    for goals in range(0, 6):
        ws.cell(row, 1, goals).alignment = ALIGN_C
        ws.cell(row, 2, round(poisson.pmf(goals, lam_a), 4)).alignment = ALIGN_C
        ws.cell(row, 3, round(poisson.pmf(goals, lam_b), 4)).alignment = ALIGN_C
        row += 1

    row += 1
    ws.cell(row, 1, "Beispiel: Ergebniswahrscheinlichkeiten 0:0 bis 5:5").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row, 1, f"{team_a} \\ {team_b}").fill = FILL_HEADER
    ws.cell(row, 1).font = FONT_HEADER
    ws.cell(row, 1).alignment = ALIGN_C
    for g_b in range(0, 6):
        hdr(ws, row, g_b + 2, g_b)
    row += 1
    for g_a in range(0, 6):
        hdr(ws, row, 1, g_a)
        for g_b in range(0, 6):
            prob = poisson.pmf(g_a, lam_a) * poisson.pmf(g_b, lam_b)
            ws.cell(row, g_b + 2, round(prob, 4)).alignment = ALIGN_C
        row += 1

    for col, width in enumerate([24, 16, 16, 70, 12, 12, 12], 1):
        cw(ws, col, width)
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Excel: Teamstaerken
# ---------------------------------------------------------------------------
def write_teamstaerken_sheet(wb):
    ws = wb.create_sheet("Teamstaerken")

    ts_path  = DATA_DIR / "team_strengths.csv"
    elo_path = DATA_DIR / "elo_checkpoint.csv"
    poi_path = DATA_DIR / "poisson_params.csv"
    mv_path  = DATA_DIR / "squad_values.csv"

    if not ts_path.exists():
        ws.cell(1, 1, "team_strengths.csv nicht gefunden."); return

    df = pd.read_csv(ts_path)
    has_mv = "squad_value_eur" in df.columns and df["squad_value_eur"].notna().any()
    has_pos_mv = (
        "market_value_attack_eur" in df.columns and
        "market_value_defense_eur" in df.columns and
        df["market_value_attack_eur"].notna().any()
    )

    # Gewichtungsinfo
    if has_mv:
        weights = "Ranking: Angriff 35% | Defensive 25% | Marktwert 30% | Elo 10%"
    else:
        weights = "Ranking ohne Marktwert: Angriff 50% | Defensive 35% | Elo 15%"

    ws.cell(1, 1, "Teamstaerken – Berechnungsgrundlage").font = Font(bold=True, size=13)
    ws.cell(2, 1, f"Gewichtung: {weights}").font = Font(italic=True, color="595959")
    ws.cell(3, 1, "Poisson: Angriff/Abwehr aus Laenderspielen ab 2015, Halbwertszeit 1,5 Jahre").font = Font(italic=True, color="595959")
    ws.cell(4, 1, "Elo: Kumulative Elo-Ratings aus allen Laenderspielen seit 2000").font    = Font(italic=True, color="595959")
    ws.cell(5, 1, "Marktwert: Kader-Gesamtmarktwert von Transfermarkt.de").font             = Font(italic=True, color="595959")

    # Header
    headers = ["Team", "Angriff", "Abwehr", "Ang.-Norm", "Def.-Norm", "Elo", "Elo-Norm"]
    if has_mv:
        headers += ["Marktwert (Mio.€)", "MV-Norm"]
    if has_pos_mv:
        headers += ["MW Angriff (Mio.€)", "MW Abwehr (Mio.€)", "MW-Ang.-Norm", "MW-Abw.-Norm"]
    headers += ["Gesamt-Stärke", "Rang"]

    for ci, h in enumerate(headers, 1):
        hdr(ws, 7, ci, h)

    # Daten sortiert nach Staerke
    df_sorted = df.sort_values("strength_combined", ascending=False).reset_index(drop=True)

    FILLS_RANK = [
        PatternFill("solid", fgColor="FFD700"),  # Gold
        PatternFill("solid", fgColor="C0C0C0"),  # Silber
        PatternFill("solid", fgColor="CD7F32"),  # Bronze
    ]

    for ri, row in df_sorted.iterrows():
        r = ri + 8
        rank = ri + 1
        vals = [
            row["team"],
            round(row.get("att", 0), 4),
            round(row.get("def", 0), 4),
            round(row.get("att_norm", 0), 3),
            round(row.get("def_quality_norm", 0), 3),
            round(row.get("elo", 1500), 1),
            round(row.get("elo_norm", 0), 3),
        ]
        if has_mv:
            mv = row.get("squad_value_eur", 0)
            vals += [round(mv / 1e6, 1) if pd.notna(mv) and mv else 0,
                     round(row.get("mv_norm", 0), 3)]
        if has_pos_mv:
            mv_att = row.get("market_value_attack_eur", 0)
            mv_def = row.get("market_value_defense_eur", 0)
            vals += [
                round(mv_att / 1e6, 1) if pd.notna(mv_att) and mv_att else 0,
                round(mv_def / 1e6, 1) if pd.notna(mv_def) and mv_def else 0,
                round(row.get("mv_att_norm", 0), 3),
                round(row.get("mv_def_norm", 0), 3),
            ]
        vals += [round(row["strength_combined"], 4), rank]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v)
            c.alignment = ALIGN_L if ci == 1 else ALIGN_C
        if rank <= 3:
            for ci in range(1, len(vals) + 1):
                ws.cell(r, ci).fill = FILLS_RANK[rank - 1]
        elif rank <= 10:
            for ci in range(1, len(vals) + 1):
                ws.cell(r, ci).fill = FILL_GROUP

    # Spaltenbreiten
    widths = [22, 10, 10, 10, 10, 10, 10, 16, 10, 17, 17, 14, 14, 14, 6]
    for i, w in enumerate(widths[:len(headers)], 1): cw(ws, i, w)

    ws.freeze_panes = "A8"


# ---------------------------------------------------------------------------
# Excel: Teams Haeufigkeiten + Prozente
# ---------------------------------------------------------------------------
def write_team_stats_sheets(wb, groups, group_pos_counts, team_stage_counts, n, as_pct=False):
    """
    Spalten: WM, Finale, HF, VF, Achtel, Sechzehntel
    Scoring:  WM=6, Finale=5, HF=4, VF=3, Achtel=2, Sechzehntel=1 -- mehr = besser
    Gruppen:  1.=4, 2.=3, 3.qual=2, 3.nq=1, 4.=0 -- mehr = besser
    """
    sheet_name = "Teams Prozente" if as_pct else "Teams Haeufigkeiten"
    ws = wb.create_sheet(sheet_name)
    all_teams = [t for g in groups.values() for t in g]

    def v(x):
        return round(x / n * 100, 1) if as_pct else int(x)

    def ko_score(t):
        sc = team_stage_counts[t]
        return (sc["p_Winner"]*6 + sc["p_F"]*5 + sc["p_SF"]*4 + sc["p_QF"]*3 +
                sc["p_R16"]*2 + sc["p_R32_entry"]*1)

    def grp_score(t):
        for grp, teams in groups.items():
            if t in teams:
                pos = group_pos_counts[grp][t]
                g3_q = team_stage_counts[t].get("p_R32_as_third", 0)
                return pos[0]*4 + pos[1]*3 + g3_q*2 + (pos[2]-g3_q)*1 + pos[3]*0
        return 0

    sorted_teams = sorted(all_teams, key=lambda t: (-ko_score(t), -grp_score(t)))

    unit = "%" if as_pct else f"/{n}"
    title = f"Teams – Simulationsergebnisse {'in Prozent' if as_pct else 'absolut'} (N={n})"
    ws.cell(1, 1, title).font = Font(bold=True, size=12)
    ws.cell(2, 1, "KO-Wertung: WM×5 + HF×4 + VF×3 + Achtel×2 + Sechzehntel×1 (mehr = besser) | Gruppen-Wertung: 1.Pl×4 + 2.Pl×3 + 3.qual×2 + 3.nq×1").font = Font(italic=True, color="595959")

    headers = [
        "Rang", "Team",
        f"Weltmeister ({unit})", f"Finale ({unit})", f"Halbfinale ({unit})", f"Viertelfinale ({unit})",
        f"Achtelfinale ({unit})", f"Sechzehntelfinale ({unit})",
        "KO-Wertung",
        f"Gr.1. ({unit})", f"Gr.2. ({unit})",
        f"Gr.3. qualif. ({unit})", f"Gr.3. nicht qual. ({unit})",
        f"Gr.4. ({unit})", "Gruppen-Wertung",
    ]
    for ci, h in enumerate(headers, 1):
        hdr(ws, 4, ci, h)

    FILL_TOP5  = PatternFill("solid", fgColor="C6EFCE")
    FILL_TOP15 = PatternFill("solid", fgColor="EBF3FB")

    for ri, team in enumerate(sorted_teams, 5):
        rank = ri - 4
        sc   = team_stage_counts[team]
        grp  = next(g for g, teams in groups.items() if team in teams)
        pos  = group_pos_counts[grp][team]
        g3_q  = sc.get("p_R32_as_third", 0)
        g3_nq = pos[2] - g3_q
        ks = ko_score(team)
        gs = grp_score(team)

        row_vals = [
            rank, team,
            v(sc["p_Winner"]), v(sc["p_F"]), v(sc["p_SF"]), v(sc["p_QF"]),
            v(sc["p_R16"]), v(sc["p_R32_entry"]),
            round(ks/n, 2) if as_pct else int(ks),
            v(pos[0]), v(pos[1]), v(g3_q), v(g3_nq), v(pos[3]),
            round(gs/n, 2) if as_pct else int(gs),
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(ri, ci, val)
            c.alignment = ALIGN_L if ci == 2 else ALIGN_C
            if rank <= 5:    c.fill = FILL_TOP5
            elif rank <= 15: c.fill = FILL_TOP15

    widths = [5, 22, 16, 14, 14, 14, 14, 18, 12, 10, 10, 18, 20, 10, 14]
    for i, w in enumerate(widths, 1): cw(ws, i, w)
    ws.freeze_panes = "C5"


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
FONT_BOLD   = Font(bold=True)
FONT_ITALIC = Font(italic=True)
FONT_FIXED  = Font(bold=True, color="1F4E79")
FILL_FIXED  = PatternFill("solid", fgColor="D6E4F0")
FILL_HEADER = PatternFill("solid", fgColor="2E75B6")
FONT_HEADER = Font(bold=True, color="FFFFFF")
FILL_GROUP  = PatternFill("solid", fgColor="EBF3FB")
FILL_WINNER = PatternFill("solid", fgColor="C6EFCE")
FILL_LOSER  = PatternFill("solid", fgColor="FFCCCC")
FILL_DRAW   = PatternFill("solid", fgColor="FFEB9C")
ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_L = Alignment(horizontal="left",   vertical="center")

def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def hdr(ws, r, c, v):
    cell = ws.cell(r, c, v); cell.font = FONT_HEADER
    cell.fill = FILL_HEADER; cell.alignment = ALIGN_C; return cell
def pct(v): return f"{v*100:.1f}%"


# ---------------------------------------------------------------------------
# Excel: Uebersicht
# ---------------------------------------------------------------------------
def write_overview(wb, groups, group_pos_counts, team_stage_counts, n):
    ws = wb.create_sheet("Uebersicht"); ws.freeze_panes = "A2"
    for ci, h in enumerate(["#","Team","16tel","Achtel","Viertel","Halb","Finale","Weltmeister"], 1):
        hdr(ws, 1, ci, h)
    all_teams = [t for g in groups.values() for t in g]
    rows = []
    for t in all_teams:
        p_r32 = sum(group_pos_counts[grp][t][pos]
                    for grp, teams in groups.items() if t in teams
                    for pos in [0, 1]) / n
        p_r32 += sum(group_pos_counts[grp][t][2]
                     for grp, teams in groups.items() if t in teams) / n * 0.67
        sc = team_stage_counts[t]
        rows.append({"team": t, "p_R32": p_r32,
                     "p_R16": sc["p_R16"]/n, "p_QF": sc["p_QF"]/n,
                     "p_SF":  sc["p_SF"]/n,  "p_F":  sc["p_F"]/n,
                     "p_Winner": sc["p_Winner"]/n})
    rows.sort(key=lambda x: -x["p_Winner"])
    for ri, row in enumerate(rows, 2):
        ws.cell(ri, 1, ri-1).alignment = ALIGN_C
        ws.cell(ri, 2, row["team"]).alignment = ALIGN_L
        for ci, col in enumerate(["p_R32","p_R16","p_QF","p_SF","p_F","p_Winner"], 3):
            c = ws.cell(ri, ci, pct(row[col])); c.alignment = ALIGN_C
            if col == "p_Winner":
                c.fill = FILL_WINNER if row[col] >= 0.08 else (FILL_GROUP if row[col] >= 0.02 else PatternFill())
    for i, w in enumerate([5,22,8,8,8,8,8,12], 1): cw(ws, i, w)


# ---------------------------------------------------------------------------
# Excel: Gruppenblatt
# ---------------------------------------------------------------------------
def write_group_sheet(wb, grp, teams, group_pos_counts, group_match_scores,
                      fixed_group_matches_flat, n):
    ws = wb.create_sheet(f"Gruppe {grp}")
    ws.cell(1, 1, f"Gruppe {grp}").font = Font(bold=True, size=12)

    pos_probs = {t: [group_pos_counts[grp][t][p] / n for p in range(4)] for t in teams}

    # Avg scores fuer diese Gruppe
    avg_scores = {k: v for k, v in group_match_scores.items() if k[0] in teams and k[1] in teams}

    # --- Tabelle 1: Konkrete Tabelle aus Ø-Ergebnissen ---
    ws.cell(2, 1, "Erwartete Tabelle").font = Font(bold=True, size=11)
    for ci, h in enumerate(["Pl.","Team","Sp","S","U","N","Tore","TD","Pkt"], 1):
        hdr(ws, 3, ci, h)
    exp_table = build_expected_table(teams, avg_scores, fixed_group_matches_flat)
    fills = [FILL_WINNER, FILL_GROUP, None, None]
    for ri, row in enumerate(exp_table, 4):
        tore = f"{row['gf']} : {row['ga']}"
        gd   = f"+{row['gd']}" if row["gd"] > 0 else str(row["gd"])
        for ci, v in enumerate([ri-3, row["team"], row["sp"], row["w"], row["d"], row["l"], tore, gd, row["pts"]], 1):
            c = ws.cell(ri, ci, v)
            c.alignment = ALIGN_L if ci == 2 else ALIGN_C
        if fills[ri-4]:
            for ci in range(1, 10): ws.cell(ri, ci).fill = fills[ri-4]

    # --- Tabelle 2: Platzierungswahrscheinlichkeiten ---
    ws.cell(10, 1, "Platzierungswahrscheinlichkeiten").font = Font(bold=True, size=11)
    for ci, h in enumerate(["Team","1. Platz","2. Platz","3. Platz","4. Platz"], 1):
        hdr(ws, 11, ci, h)
    for ri, team in enumerate(sorted(teams, key=lambda t: -pos_probs[t][0]), 12):
        ws.cell(ri, 1, team).alignment = ALIGN_L
        for ci, pos in enumerate(range(4), 2):
            c = ws.cell(ri, ci, pct(pos_probs[team][pos])); c.alignment = ALIGN_C
            if pos == 0 and pos_probs[team][0] >= 0.5: c.fill = FILL_WINNER

    # --- Spiele ---
    ws.cell(18, 1, "Spiele").font = Font(bold=True, size=11)
    for ci, h in enumerate(["Heim","Ergebnis","Auswärts","Heimsieg%","Unentsch.%","Ausw-Sieg%"], 1):
        hdr(ws, 19, ci, h)
    row_n = 20
    for (h, a), s in sorted(avg_scores.items()):
        total = s["wins_h"] + s["draws"] + s["wins_a"]
        p_wh = s["wins_h"] / max(total, 1)
        p_d  = s["draws"]  / max(total, 1)
        p_wa = s["wins_a"] / max(total, 1)
        if (h, a) in fixed_group_matches_flat:
            gh, ga = fixed_group_matches_flat[(h, a)]
            score_str = f"{gh} : {ga}"; font = FONT_FIXED; fill = FILL_FIXED
        elif (a, h) in fixed_group_matches_flat:
            ga, gh = fixed_group_matches_flat[(a, h)]
            score_str = f"{gh} : {ga}"; font = FONT_FIXED; fill = FILL_FIXED
        else:
            avg_h = np.mean(s["goals_h"]) if s["goals_h"] else 0
            avg_a = np.mean(s["goals_a"]) if s["goals_a"] else 0
            gh, ga = score_from_avg(avg_h, avg_a, p_wh, p_wa)
            score_str = f"{gh} : {ga}"; font = FONT_ITALIC; fill = None
        for ci, v in enumerate([h, score_str, a, pct(p_wh), pct(p_d), pct(p_wa)], 1):
            c = ws.cell(row_n, ci, v)
            c.font = font; c.alignment = ALIGN_L if ci in [1,3] else ALIGN_C
            if fill: c.fill = fill
        row_n += 1

    for i, w in enumerate([22,14,22,10,10,10], 1): cw(ws, i, w)


def build_table_from_scores(teams, scores):
    records = {t: {"sp":0,"w":0,"d":0,"l":0,"gf":0,"ga":0,"pts":0} for t in teams}
    for (h, a), (gh, ga) in scores.items():
        if h not in records or a not in records:
            continue
        records[h]["sp"] += 1; records[a]["sp"] += 1
        records[h]["gf"] += gh; records[h]["ga"] += ga
        records[a]["gf"] += ga; records[a]["ga"] += gh
        if gh > ga:
            records[h]["w"] += 1; records[h]["pts"] += 3; records[a]["l"] += 1
        elif gh == ga:
            records[h]["d"] += 1; records[h]["pts"] += 1
            records[a]["d"] += 1; records[a]["pts"] += 1
        else:
            records[a]["w"] += 1; records[a]["pts"] += 3; records[h]["l"] += 1
    table = [{"team": t, "gd": r["gf"] - r["ga"], **r} for t, r in records.items()]
    table.sort(key=lambda x: (-x["pts"], -x["gd"], -x["gf"]))
    return table


def write_groups_sheet(wb, groups, representative):
    """Alle Gruppen kompakt in einem Blatt, jeweils vier nebeneinander."""
    ws = wb.create_sheet("Gruppen")
    ws.cell(1, 1, "Gruppenphase - repraesentativer simulierter Turnierverlauf").font = Font(bold=True, size=13)
    ws.cell(2, 1, f"Turnier-ID: {representative['tournament_id']}  |  "
                  f"Matching-Score: {representative['matching_score']:.3f}").font = FONT_ITALIC
    headers = ["Pl.", "Team", "Sp", "S", "U", "N", "Tore", "TD", "Pkt"]
    fills = [FILL_WINNER, FILL_GROUP, None, None]

    for idx, grp in enumerate(sorted(groups)):
        block_row = 4 + (idx // 4) * 18
        block_col = 1 + (idx % 4) * 10
        teams = groups[grp]
        scores = {k: v for k, v in representative["group_scores"].items()
                  if k[0] in teams and k[1] in teams}
        table = build_table_from_scores(teams, scores)

        ws.cell(block_row, block_col, f"Gruppe {grp}").font = Font(bold=True, size=12)
        for offset, header in enumerate(headers):
            hdr(ws, block_row + 1, block_col + offset, header)
        for pos, row in enumerate(table, 1):
            tore = f"{row['gf']} : {row['ga']}"
            gd = f"+{row['gd']}" if row["gd"] > 0 else str(row["gd"])
            vals = [pos, row["team"], row["sp"], row["w"], row["d"], row["l"], tore, gd, row["pts"]]
            for offset, value in enumerate(vals):
                cell = ws.cell(block_row + 1 + pos, block_col + offset, value)
                cell.alignment = ALIGN_L if offset == 1 else ALIGN_C
                if fills[pos - 1]:
                    cell.fill = fills[pos - 1]

        ws.cell(block_row + 7, block_col, "Spiele").font = FONT_BOLD
        for offset, header in enumerate(["Heim", "Ergebnis", "Auswaerts"]):
            hdr(ws, block_row + 8, block_col + offset, header)
        for row_idx, ((h, a), (gh, ga)) in enumerate(sorted(scores.items()), block_row + 9):
            ws.cell(row_idx, block_col, h).alignment = ALIGN_L
            ws.cell(row_idx, block_col + 1, f"{gh} : {ga}").alignment = ALIGN_C
            ws.cell(row_idx, block_col + 2, a).alignment = ALIGN_L

    for block_col in [1, 11, 21, 31]:
        for offset, width in enumerate([5, 22, 5, 5, 5, 5, 10, 7, 6]):
            cw(ws, block_col + offset, width)
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Excel: Turnierbaum
# ---------------------------------------------------------------------------
def write_bracket_sheet(wb, ko_results_summary, fixed_ko, representative_ko_scores=None,
                        representative_ko_results=None):
    ws = wb.create_sheet("Turnierbaum")

    # Spalten-Header (korrekte Bezeichnungen)
    round_labels = {
        1:  "Sechzehntelfinale (32 Teams)",
        4:  "Achtelfinale (16 Teams)",
        7:  "Viertelfinale (8 Teams)",
        10: "Halbfinale (4 Teams)",
        13: "Finale",
    }
    for col, label in round_labels.items():
        c = ws.cell(1, col, label); c.font = FONT_HEADER; c.fill = FILL_HEADER
        c.alignment = ALIGN_C
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)

    # Slot-Beschreibung links neben R32-Matches (in Spalte 0 = praktisch Spalte A, Teams in B)
    # Aber R32 Teams stehen in col 1 und Score in col 2 - wir brauchen eine extra Spalte
    # Wir verschieben alles: Beschreibung in col 1, Team in col 2, Score in col 3
    # und passen BRACKET-Positionen an mit +1 fuer col
    # STATTDESSEN: Beschreibung in Kommentar oder extra Spalte
    # Einfachere Loesung: Beschreibung direkt als zweite Zeile im Bracket-Layout einfuegen

    for mnr, (r1, r2, ct, cs) in BRACKET.items():
        r_stat = ko_results_summary.get(mnr)  # Statistiken (Wahrscheinlichkeiten)
        is_fixed = mnr in fixed_ko

        if is_fixed:
            fk = fixed_ko[mnr]
            ta, tb = fk["team_a"], fk["team_b"]
            ga, gb = fk["goals_a"], fk["goals_b"]
            score_str = f"{ga} : {gb}"
            if fk.get("extra"): score_str += f" {fk['extra']}"
            font_a = FONT_FIXED if ga >  gb else Font(bold=True, color="999999")
            font_b = FONT_FIXED if gb >= ga else Font(bold=True, color="999999")
            fill_a = FILL_WINNER if ga > gb else FILL_LOSER
            fill_b = FILL_WINNER if gb > ga else FILL_LOSER
            if ga == gb: fill_a = fill_b = FILL_DRAW
        else:
            # Teams und Ergebnisse aus einem real simulierten Referenzpfad.
            if representative_ko_scores and mnr in representative_ko_scores:
                ta, tb, ga, gb, extra = representative_ko_scores[mnr]
            elif r_stat:
                ta, tb = r_stat["team_a"], r_stat["team_b"]
                ga, gb = r_stat["score_a"], r_stat["score_b"]
                extra = ""
            else:
                continue
            score_str = f"{ga} : {gb}"
            if extra:
                score_str += f" {extra}"
            winner = (representative_ko_results or {}).get(mnr)
            font_a = Font(italic=True, bold=winner == ta, color="1F4E79" if winner == ta else "000000")
            font_b = Font(italic=True, bold=winner == tb, color="1F4E79" if winner == tb else "000000")
            fill_a = FILL_WINNER if winner == ta else FILL_LOSER
            fill_b = FILL_WINNER if winner == tb else FILL_LOSER

        c1 = ws.cell(r1+1, ct, ta); c1.font = font_a; c1.alignment = ALIGN_L
        c2 = ws.cell(r2+1, ct, tb); c2.font = font_b; c2.alignment = ALIGN_L
        cs_c = ws.cell(r1+1, cs, score_str)
        cs_c.font = FONT_FIXED if is_fixed else FONT_ITALIC; cs_c.alignment = ALIGN_C
        if fill_a: c1.fill = fill_a
        if fill_b: c2.fill = fill_b
        if is_fixed: cs_c.fill = FILL_FIXED

    for col in [1,4,7,10,13]: cw(ws, col, 22); cw(ws, col+1, 10)
    for col in [3,6,9,12]:    cw(ws, col, 2)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Excel: Tipps-Blatt
# ---------------------------------------------------------------------------
def load_previous_tipps(path):
    """Liest die Tipps aus der bestehenden Excel, bevor sie ueberschrieben wird."""
    previous = {"group": {}, "ko": {}}
    if not path.exists():
        return previous

    try:
        old_wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if "Tipps" not in old_wb.sheetnames:
            old_wb.close()
            return previous

        ws = old_wb["Tipps"]
        mode = None
        cols = {}
        for row in ws.iter_rows():
            values = [cell.value for cell in row]
            headers = {
                str(value).strip(): idx
                for idx, value in enumerate(values)
                if value is not None
            }

            if "Datum" in headers and "Heim" in headers and "Tipp" in headers:
                away_col = headers.get("Auswärts", headers.get("Auswaerts"))
                if away_col is not None:
                    mode = "group"
                    cols = {
                        "home": headers["Heim"],
                        "tip": headers["Tipp"],
                        "away": away_col,
                    }
                continue

            if "Match" in headers and "Tipp" in headers:
                away_col = headers.get("Auswärts (wahrscheinlich)", headers.get("Auswaerts (wahrscheinlich)"))
                if away_col is not None:
                    mode = "ko"
                    cols = {
                        "match": headers["Match"],
                        "tip": headers["Tipp"],
                    }
                continue

            if mode == "group" and cols:
                home = values[cols["home"]] if cols["home"] < len(values) else None
                away = values[cols["away"]] if cols["away"] < len(values) else None
                tip = values[cols["tip"]] if cols["tip"] < len(values) else None
                if home and away and tip:
                    previous["group"][(str(home), str(away))] = str(tip)

            elif mode == "ko" and cols:
                match_value = values[cols["match"]] if cols["match"] < len(values) else None
                tip = values[cols["tip"]] if cols["tip"] < len(values) else None
                if match_value and tip:
                    try:
                        mnr = int(str(match_value).split()[-1])
                    except ValueError:
                        continue
                    previous["ko"][mnr] = str(tip)

        old_wb.close()
    except Exception as exc:
        print(f"Vorherige Tipps konnten nicht gelesen werden: {exc}")

    return previous


def write_tipps_sheet(wb, groups, group_pos_counts, team_stage_counts,
                      group_match_scores, fixed_group_matches, ko_results_summary,
                      fixed_ko, n, previous_tipps=None):
    ws = wb.create_sheet("Tipps")
    ws.freeze_panes = "A2"
    previous_tipps = previous_tipps or {"group": {}, "ko": {}}

    FILL_Q    = PatternFill("solid", fgColor="FFF2CC")   # Fragen
    FILL_GRP  = PatternFill("solid", fgColor="DDEBF7")   # Gruppenspiele
    FILL_KO   = PatternFill("solid", fgColor="E2EFDA")   # KO-Spiele
    FILL_CONF = PatternFill("solid", fgColor="C6EFCE")   # Sicherer Tipp (>60%)
    FILL_CHANGED = PatternFill("solid", fgColor="FCE4D6") # Tipp hat sich geaendert
    FONT_BIG  = Font(bold=True, size=13)
    FONT_MED  = Font(bold=True, size=11)

    row = 1

    # ---- Abschnitt 1: Vorher-Fragen ----
    ws.cell(row, 1, "VORHER ZU BEANTWORTENDE FRAGEN").font = FONT_BIG
    row += 1
    for ci, h in enumerate(["Frage", "Empfehlung", "Wahrscheinlichkeit", "Hinweis"], 1):
        c = ws.cell(row, ci, h); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C
    row += 1

    all_teams = [t for g in groups.values() for t in g]

    # Gruppensieg-Fragen
    for grp in sorted(groups.keys()):
        teams = groups[grp]
        winner = max(teams, key=lambda t: group_pos_counts[grp][t][0])
        p_win  = group_pos_counts[grp][winner][0] / n
        second = max([t for t in teams if t != winner], key=lambda t: group_pos_counts[grp][t][0])
        p2     = group_pos_counts[grp][second][0] / n
        hint   = f"Zweitplatzierter wahrscheinlich: {second} ({pct(p2)})"
        ws.cell(row, 1, f"Wer gewinnt die Gruppe {grp}?").alignment = ALIGN_L
        ws.cell(row, 2, winner).font = Font(bold=True); ws.cell(row, 2).alignment = ALIGN_L
        ws.cell(row, 3, pct(p_win)).alignment = ALIGN_C
        ws.cell(row, 4, hint).font = Font(italic=True); ws.cell(row, 4).alignment = ALIGN_L
        for ci in range(1, 5): ws.cell(row, ci).fill = FILL_Q
        if p_win >= 0.50:
            ws.cell(row, 3).fill = FILL_CONF
        row += 1

    row += 1  # Leerzeile

    # Halbfinale (4 Teams mit hoechster p_SF)
    sf_teams = sorted(all_teams, key=lambda t: -team_stage_counts[t]["p_SF"])[:4]
    sf_str   = " / ".join(sf_teams)
    sf_probs = " / ".join([pct(team_stage_counts[t]["p_SF"] / n) for t in sf_teams])
    ws.cell(row, 1, "Wer erreicht das Halbfinale?").alignment = ALIGN_L
    ws.cell(row, 2, sf_str).font = Font(bold=True); ws.cell(row, 2).alignment = ALIGN_L
    ws.cell(row, 3, sf_probs).font = Font(italic=True); ws.cell(row, 3).alignment = ALIGN_C
    ws.cell(row, 4, "4 Teams mit hoechster Halbfinale-Wahrscheinlichkeit").font = Font(italic=True)
    for ci in range(1, 5): ws.cell(row, ci).fill = FILL_Q
    row += 1

    # Weltmeister
    wm_winner = max(all_teams, key=lambda t: team_stage_counts[t]["p_Winner"])
    p_wm      = team_stage_counts[wm_winner]["p_Winner"] / n
    top3_wm   = sorted(all_teams, key=lambda t: -team_stage_counts[t]["p_Winner"])[:3]
    hint_wm   = "Top 3: " + ", ".join([f"{t} ({pct(team_stage_counts[t]['p_Winner']/n)})" for t in top3_wm])
    ws.cell(row, 1, "Wer wird Weltmeister?").alignment = ALIGN_L
    ws.cell(row, 2, wm_winner).font = Font(bold=True); ws.cell(row, 2).alignment = ALIGN_L
    ws.cell(row, 3, pct(p_wm)).alignment = ALIGN_C
    ws.cell(row, 4, hint_wm).font = Font(italic=True); ws.cell(row, 4).alignment = ALIGN_L
    for ci in range(1, 5): ws.cell(row, ci).fill = FILL_Q
    row += 1

    # Torschuetze (ohne Spieler-Daten: Team mit meisten Gesamttoren)
    ws.cell(row, 1, "Welche Mannschaft: Spieler mit meisten Toren?").alignment = ALIGN_L
    ws.cell(row, 2, wm_winner).font = Font(bold=True); ws.cell(row, 2).alignment = ALIGN_L
    ws.cell(row, 3, "Schätzung").alignment = ALIGN_C
    ws.cell(row, 4, "Kein Spieler-Modell – Empfehlung: Favorit-Team (meiste Tore erwartet)").font = Font(italic=True)
    for ci in range(1, 5): ws.cell(row, ci).fill = FILL_Q
    row += 2

    # ---- Abschnitt 2: Gruppenspiele ----
    ws.cell(row, 1, "GRUPPENSPIELE – SPIELTIPPS").font = FONT_MED
    row += 1
    for ci, h in enumerate(["Datum","Heim","Tipp","Vorheriger Tipp","Auswärts","Heimsieg%","Unentsch.%","Auswärtssieg%","Konfidenz"], 1):
        c = ws.cell(row, ci, h); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C
    row += 1

    # Lade Spielplan
    matches_path = DATA_DIR / "wm2026_matches_group.csv"
    if matches_path.exists():
        matches_df = pd.read_csv(matches_path, parse_dates=["date"])
        matches_df = matches_df.sort_values("date")
        for _, m in matches_df.iterrows():
            h, a = m["team_home"], m["team_away"]
            grp  = m["group"]
            fixed_flat = fixed_group_matches.get(grp, {})

            # Bereits gespielt?
            if (h, a) in fixed_flat:
                gh, ga = fixed_flat[(h, a)]
                score_str = f"{gh} : {ga}"
                p_wh = p_d = p_wa = "—"
                konf = "✓ Gespielt"
                fill = FILL_FIXED if 'FILL_FIXED' in dir() else FILL_GROUP
            else:
                key = (h, a) if (h, a) in group_match_scores else (a, h)
                s = group_match_scores.get(key, {})
                if s:
                    tot = s["wins_h"] + s["draws"] + s["wins_a"]
                    p_wh_f = s["wins_h"] / max(tot, 1)
                    p_d_f  = s["draws"]  / max(tot, 1)
                    p_wa_f = s["wins_a"] / max(tot, 1)
                    if key != (h, a):
                        p_wh_f, p_wa_f = p_wa_f, p_wh_f
                    if key == (h, a):
                        avg_h = np.mean(s["goals_h"]) if s.get("goals_h") else 0
                        avg_a = np.mean(s["goals_a"]) if s.get("goals_a") else 0
                        score_counts = s.get("score_counts", {})
                    else:
                        avg_h = np.mean(s["goals_a"]) if s.get("goals_a") else 0
                        avg_a = np.mean(s["goals_h"]) if s.get("goals_h") else 0
                        score_counts = {
                            (score_a, score_h): count
                            for (score_h, score_a), count in s.get("score_counts", {}).items()
                        }
                    (gh, ga), score_agrees = select_tip_score(avg_h, avg_a, p_wh_f, p_wa_f, score_counts)
                    score_str = f"{gh} : {ga}"
                    p_wh = pct(p_wh_f); p_d = pct(p_d_f); p_wa = pct(p_wa_f)
                    max_p = max(p_wh_f, p_d_f, p_wa_f)
                    konf  = "★★★" if max_p > 0.65 else ("★★" if max_p > 0.50 else "★")
                    if score_agrees:
                        konf += "!"
                    fill  = FILL_CONF if max_p > 0.65 else None
                else:
                    score_str = "—"; p_wh = p_d = p_wa = "—"; konf = "—"; fill = None

            date_str = m["date"].strftime("%d.%m.%y") if pd.notna(m["date"]) else "—"
            previous_score = previous_tipps["group"].get((h, a), "")
            changed = previous_score and previous_score != score_str
            for ci, v in enumerate([date_str, h, score_str, previous_score, a, p_wh, p_d, p_wa, konf], 1):
                c = ws.cell(row, ci, v)
                c.alignment = ALIGN_L if ci in [2, 5] else ALIGN_C
                if ci == 3: c.font = Font(bold=True)
                if fill: c.fill = fill
                if changed and ci in [3, 4]:
                    c.fill = FILL_CHANGED
            row += 1

    row += 1

    # ---- Abschnitt 3: KO-Spiele ----
    ws.cell(row, 1, "KO-SPIELE – SPIELTIPPS (basierend auf wahrscheinlichstem Turnierverlauf)").font = FONT_MED
    row += 1
    for ci, h in enumerate(["Match","Runde","Heim (wahrscheinlich)","Tipp","Vorheriger Tipp","Auswärts (wahrscheinlich)","Sieger%","Konf."], 1):
        c = ws.cell(row, ci, h); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C
    row += 1

    round_names = {
        **{mnr: "Sechzehntelfinale" for mnr in range(73, 89)},
        **{mnr: "Achtelfinale"      for mnr in range(89, 97)},
        **{mnr: "Viertelfinale"     for mnr in range(97, 101)},
        101: "Halbfinale", 102: "Halbfinale",
        103: "Spiel um Platz 3", 104: "Finale",
    }

    for mnr in sorted(ko_results_summary.keys()):
        r = ko_results_summary[mnr]
        ta, tb = r["team_a"], r["team_b"]
        p_a = r["p_a_wins"]
        ga, gb = r["score_a"], r["score_b"]
        score_str = f"{ga} : {gb}"
        winner = ta if p_a >= 0.5 else tb
        konf = "★★★" if max(p_a, 1-p_a) > 0.65 else ("★★" if max(p_a, 1-p_a) > 0.55 else "★")
        if r.get("score_agrees_with_avg"):
            konf += "!"
        fill = FILL_CONF if max(p_a, 1-p_a) > 0.65 else (FILL_KO if max(p_a, 1-p_a) > 0.55 else None)
        previous_score = previous_tipps["ko"].get(mnr, "")
        changed = previous_score and previous_score != score_str

        for ci, v in enumerate([f"Match {mnr}", round_names.get(mnr,"KO"), ta, score_str, previous_score, tb, pct(max(p_a,1-p_a)), konf], 1):
            c = ws.cell(row, ci, v)
            c.alignment = ALIGN_L if ci in [3, 6] else ALIGN_C
            if ci == 4: c.font = Font(bold=True)
            if fill: c.fill = fill
            if changed and ci in [4, 5]:
                c.fill = FILL_CHANGED
        row += 1

    # Spaltenbreiten
    for i, w in enumerate([12, 20, 28, 10, 14, 28, 10, 8, 8], 1): cw(ws, i, w)


# ---------------------------------------------------------------------------
# Excel: Drittplatzierte
# ---------------------------------------------------------------------------
def write_thirds_sheet(wb, groups, group_pos_counts, group_match_scores,
                       fixed_group_matches, team_stage_counts, n):
    """
    Drittplatzierte – konsistent mit dem Turnierbaum:
    - Wahrscheinlichster Dritter per Gruppe aus group_pos_counts (Simulation)
    - Qualifikationsstatus aus team_stage_counts['p_R32_as_third']
    - Tabellen-Stats aus Ø-Ergebnissen
    """
    ws = wb.create_sheet("Drittplatzierte")
    ws.cell(1, 1, "Drittplatzierte – Qualifikationsrangliste").font = Font(bold=True, size=12)
    ws.cell(2, 1, "Wahrscheinlichster Dritter je Gruppe (aus 1000 Simulationen). "
                  "Qualifikationsstatus: wie oft war dieses Team unter den besten 8 Dritten?").font = Font(italic=True, color="595959")

    thirds = []
    for grp, teams in groups.items():
        # Wahrscheinlichster Dritter aus Simulation (gleiche Methode wie Turnierbaum)
        expected_pos = {}
        for t in teams:
            counts = group_pos_counts[grp][t]
            total  = max(sum(counts), 1)
            expected_pos[t] = sum(pos * cnt for pos, cnt in enumerate(counts)) / total
        ranked = sorted(teams, key=lambda t: expected_pos[t])
        t3 = ranked[2] if len(ranked) >= 3 else ranked[-1]

        # Stats aus Ø-Ergebnissen
        fixed_flat = fixed_group_matches.get(grp, {})
        avg_scores = {k: v for k, v in group_match_scores.items()
                      if k[0] in teams and k[1] in teams}
        exp_table  = build_expected_table(teams, avg_scores, fixed_flat)
        stats      = next((r for r in exp_table if r["team"] == t3), {})

        # Qualifikationsrate aus Simulation
        times_third    = group_pos_counts[grp][t3][2]
        times_qual     = team_stage_counts[t3].get("p_R32_as_third", 0)
        qual_rate      = times_qual / max(times_third, 1)

        thirds.append({
            "grp": grp, "team": t3,
            "times_third": times_third,
            "times_qual":  times_qual,
            "qual_rate":   qual_rate,
            "pts": stats.get("pts", 0), "gf": stats.get("gf", 0),
            "ga":  stats.get("ga",  0), "gd": stats.get("gd", 0),
            "sp":  stats.get("sp",  3), "w":  stats.get("w",  0),
            "d":   stats.get("d",   0), "l":  stats.get("l",  0),
        })

    # Sortierung: Qualifikationsrate (wie oft unter besten 8)
    thirds.sort(key=lambda x: (-x["qual_rate"], -x["pts"], -x["gd"], -x["gf"]))

    for ci, h in enumerate(["Rang","Gr.","Team","Sp","S","U","N","Tore","TD","Pkt",
                              f"3.Platz (/{n})","davon qualif.","Qualif.-%","Status"], 1):
        hdr(ws, 4, ci, h)

    for ri, row in enumerate(thirds, 5):
        rank   = ri - 4
        tore   = f"{row['gf']} : {row['ga']}"
        gd_str = f"+{row['gd']}" if row["gd"] > 0 else str(row["gd"])
        status = "✓ Qualifiziert" if row["qual_rate"] >= 0.5 else "✗ Ausgeschieden"
        q_pct  = f"{row['qual_rate']*100:.1f}%"
        vals   = [rank, row["grp"], row["team"],
                  row["sp"], row["w"], row["d"], row["l"],
                  tore, gd_str, row["pts"],
                  row["times_third"], row["times_qual"], q_pct, status]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            c.alignment = ALIGN_L if ci == 3 else ALIGN_C
        fill = FILL_WINNER if row["qual_rate"] >= 0.5 else FILL_LOSER
        for ci in range(1, 15): ws.cell(ri, ci).fill = fill

    for i, w in enumerate([5,5,22,5,5,5,5,10,7,6,12,14,10,15], 1): cw(ws, i, w)


def write_representative_thirds_sheet(wb, groups, representative,
                                      group_pos_counts, team_stage_counts, n):
    """Drittenrangliste aus demselben simulierten Verlauf wie der Turnierbaum."""
    ws = wb.create_sheet("Drittplatzierte")
    ws.cell(1, 1, "Drittplatzierte - repraesentativer simulierter Turnierverlauf").font = Font(bold=True, size=12)
    ws.cell(2, 1, "Die ersten acht Teams sind in genau diesem Verlauf qualifiziert. "
                  "Die Quote rechts stammt aus allen Simulationen.").font = FONT_ITALIC

    qualified_thirds = {
        team for slot, team in representative["qualifiers"].items() if slot.startswith("3rd Group ")
    }
    thirds = []
    for grp, teams in groups.items():
        ranking, records = representative["group_standings"][grp]
        team = ranking[2]
        scores = {k: v for k, v in representative["group_scores"].items()
                  if k[0] in teams and k[1] in teams}
        stats = next(row for row in build_table_from_scores(teams, scores) if row["team"] == team)
        times_third = group_pos_counts[grp][team][2]
        times_qual = team_stage_counts[team].get("p_R32_as_third", 0)
        thirds.append({
            "grp": grp, "team": team,
            "qualified": team in qualified_thirds,
            "times_third": times_third,
            "times_qual": times_qual,
            "qual_rate": times_qual / max(times_third, 1),
            **stats,
        })

    thirds.sort(key=lambda row: (-row["pts"], -row["gd"], -row["gf"]))
    headers = ["Rang", "Gr.", "Team", "Sp", "S", "U", "N", "Tore", "TD", "Pkt",
               f"3.Platz (/{n})", "davon qualif.", "Qualif.-%", "Status"]
    for ci, header in enumerate(headers, 1):
        hdr(ws, 4, ci, header)

    for ri, row in enumerate(thirds, 5):
        rank = ri - 4
        tore = f"{row['gf']} : {row['ga']}"
        gd = f"+{row['gd']}" if row["gd"] > 0 else str(row["gd"])
        status = "Qualifiziert" if row["qualified"] else "Ausgeschieden"
        vals = [rank, row["grp"], row["team"], row["sp"], row["w"], row["d"], row["l"],
                tore, gd, row["pts"], row["times_third"], row["times_qual"],
                f"{row['qual_rate']*100:.1f}%", status]
        for ci, value in enumerate(vals, 1):
            cell = ws.cell(ri, ci, value)
            cell.alignment = ALIGN_L if ci == 3 else ALIGN_C
            cell.fill = FILL_WINNER if row["qualified"] else FILL_LOSER

    for i, width in enumerate([5,5,22,5,5,5,5,10,7,6,12,14,10,15], 1):
        cw(ws, i, width)


def write_reference_variants_sheet(wb, ranked_paths, team_stage_counts, n, limit=10):
    """Top-Simulationslaeufe nach Matching-Score mit Podium."""
    ws = wb.create_sheet("Referenzvarianten")
    ws.cell(1, 1, "Top-Referenzvarianten aus den simulierten Turnierverlaeufen").font = Font(bold=True, size=12)
    ws.cell(2, 1, "Rang 1 wird fuer Gruppen, Drittplatzierte und Turnierbaum verwendet. "
                  "Der Score bewertet den gesamten Turnierverlauf.").font = FONT_ITALIC
    headers = ["Rang", "Turnier-ID", "Matching-Score", "Weltmeister", "WM-Quote",
               "2. Platz", "3. Platz", "4. Platz"]
    for ci, header in enumerate(headers, 1):
        hdr(ws, 4, ci, header)

    for ri, path in enumerate(ranked_paths[:limit], 5):
        champion, runner_up, third, fourth = get_podium(path)
        values = [ri - 4, path["tournament_id"], round(path["matching_score"], 3),
                  champion, pct(team_stage_counts[champion]["p_Winner"] / n),
                  runner_up, third, fourth]
        for ci, value in enumerate(values, 1):
            cell = ws.cell(ri, ci, value)
            cell.alignment = ALIGN_L if ci in [4, 6, 7, 8] else ALIGN_C
            if ri == 5:
                cell.fill = FILL_WINNER

    for i, width in enumerate([6, 12, 16, 22, 10, 22, 22, 22], 1):
        cw(ws, i, width)


# ---------------------------------------------------------------------------
# Tipp-Persistenz
# ---------------------------------------------------------------------------
def save_tips_to_history(group_match_scores, fixed_group_matches, matches_path, ko_results_summary, fixed_ko):
    """
    Schreibt aktuelle Tipps in data/tips_history.csv.
    Bereits gespielte Spiele werden NICHT ueberschrieben (Tipp gilt als abgegeben).
    Noch nicht gespielte Spiele werden aktualisiert.
    """
    history_path = DATA_DIR / "tips_history.csv"
    fieldnames = ["match_nr", "home", "away", "tip_home", "tip_away", "locked", "run_timestamp"]

    # Bestehende Tipps laden (key: match_nr als String)
    existing = {}
    if history_path.exists():
        with open(history_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                existing[row["match_nr"]] = row

    now = datetime.now().isoformat(timespec="seconds")

    # Gruppenspiele
    if matches_path.exists():
        matches_df = pd.read_csv(matches_path, parse_dates=["date"])
        for _, m in matches_df.iterrows():
            key = str(int(m["match_nr"]))
            grp = m["group"]
            h, a = m["team_home"], m["team_away"]
            fixed_flat = fixed_group_matches.get(grp, {})
            already_played = (h, a) in fixed_flat or (a, h) in fixed_flat

            if already_played and key in existing:
                # Eingefroren: nicht aendern
                continue

            if already_played:
                # Spiel gespielt, kein vorheriger Tipp gespeichert → nichts einfrieren
                continue

            # Noch nicht gespielt: aktuellen Tipp berechnen und updaten
            score_key = (h, a) if (h, a) in group_match_scores else (a, h)
            s = group_match_scores.get(score_key, {})
            if not s:
                continue
            tot = s["wins_h"] + s["draws"] + s["wins_a"]
            if score_key == (h, a):
                p_wh = s["wins_h"] / max(tot, 1)
                p_wa = s["wins_a"] / max(tot, 1)
                avg_h = np.mean(s["goals_h"]) if s.get("goals_h") else 0
                avg_a = np.mean(s["goals_a"]) if s.get("goals_a") else 0
                sc = s.get("score_counts", {})
            else:
                p_wh = s["wins_a"] / max(tot, 1)
                p_wa = s["wins_h"] / max(tot, 1)
                avg_h = np.mean(s["goals_a"]) if s.get("goals_a") else 0
                avg_a = np.mean(s["goals_h"]) if s.get("goals_h") else 0
                sc = {(ga, gh): c for (gh, ga), c in s.get("score_counts", {}).items()}
            (tip_h, tip_a), _ = select_tip_score(avg_h, avg_a, p_wh, p_wa, sc)
            existing[key] = {
                "match_nr": key, "home": h, "away": a,
                "tip_home": str(tip_h), "tip_away": str(tip_a),
                "locked": "False", "run_timestamp": now,
            }

    # KO-Spiele
    for mnr, r in ko_results_summary.items():
        key = str(mnr)
        already_played = mnr in fixed_ko
        if already_played and key in existing:
            continue
        if already_played:
            # KO-Spiel gespielt, kein vorheriger Tipp → nichts einfrieren
            continue
        ta, tb = r["team_a"], r["team_b"]
        tip_h, tip_a = r["score_a"], r["score_b"]
        existing[key] = {
            "match_nr": key, "home": ta, "away": tb,
            "tip_home": str(tip_h), "tip_away": str(tip_a),
            "locked": "False", "run_timestamp": now,
        }

    # Sortiert nach match_nr schreiben
    rows = sorted(existing.values(), key=lambda r: int(r["match_nr"]))
    with open(history_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Tipps gespeichert: {history_path.name} ({len(rows)} Spiele)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Lade Eingaben ...")
    strengths, groups, ko_df = load_inputs()
    fixed_group_matches = load_fixed_group_matches()
    fixed_ko            = load_fixed_ko_matches()
    out = Path(__file__).resolve().parent / "wm2026_simulation.xlsx"
    previous_tipps = load_previous_tipps(out)

    print(f"Simuliere {N_TOURNAMENTS} vollstaendige Turniere ...")
    print("(jedes Team erscheint pro Simulation genau einmal)")

    group_pos_counts, group_match_scores, ko_match_stats, team_stage_counts, tournament_paths, n = run_n_tournaments(
        groups, strengths, fixed_group_matches, ko_df, fixed_ko, N_TOURNAMENTS
    )

    ko_results_summary = get_ko_result_summary(ko_match_stats, n)

    print("Speichere Tipps ...")
    matches_path = DATA_DIR / "wm2026_matches_group.csv"
    save_tips_to_history(
        group_match_scores, fixed_group_matches, matches_path,
        ko_results_summary, fixed_ko,
    )

    print("Waehle repraesentativen simulierten Turnierverlauf ...")
    ranked_paths = rank_representative_tournaments(
        tournament_paths, group_pos_counts, team_stage_counts, n
    )
    representative = ranked_paths[0]
    print(f"  Turnier-ID: {representative['tournament_id']} | "
          f"Matching-Score: {representative['matching_score']:.3f}")

    print("\nTop 10 Referenzvarianten:")
    print(f"{'Rang':>4} {'ID':>5} {'Score':>8}  {'Weltmeister':<20} {'2. Platz':<20} {'3. Platz':<20} {'4. Platz':<20}")
    for rank, path in enumerate(ranked_paths[:10], 1):
        champion, runner_up, third, fourth = get_podium(path)
        print(f"{rank:>4} {path['tournament_id']:>5} {path['matching_score']:>8.3f}  "
              f"{champion:<20} {runner_up:<20} {third:<20} {fourth:<20}")

    print("\nSchreibe Excel ...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Reihenfolge der Blaetter:
    # 1. Erklaerung, 2. Tipps, 3. Teamstaerken, 4. Teams Haeufigkeiten,
    # 5. Teams Prozente, 6. Referenzvarianten, 7. Gruppen, 8. Turnierbaum,
    # 9. Drittplatzierte
    write_explanation_sheet(wb, strengths)
    write_tipps_sheet(wb, groups, group_pos_counts, team_stage_counts,
                      group_match_scores, fixed_group_matches,
                      ko_results_summary, fixed_ko, n, previous_tipps)
    write_teamstaerken_sheet(wb)
    write_team_stats_sheets(wb, groups, group_pos_counts, team_stage_counts, n, as_pct=False)
    write_team_stats_sheets(wb, groups, group_pos_counts, team_stage_counts, n, as_pct=True)
    write_reference_variants_sheet(wb, ranked_paths, team_stage_counts, n)
    write_groups_sheet(wb, groups, representative)
    write_bracket_sheet(
        wb, ko_results_summary, fixed_ko,
        representative["ko_scores"], representative["ko_results"]
    )
    write_representative_thirds_sheet(
        wb, groups, representative, group_pos_counts, team_stage_counts, n
    )

    wb.save(out)
    print(f"Gespeichert: {out}")

    # Top 10 ausgeben
    all_teams = [t for g in groups.values() for t in g]
    top = sorted(all_teams, key=lambda t: -team_stage_counts[t]["p_Winner"])[:10]
    print("\nTop 10 Weltmeister-Wahrscheinlichkeit:")
    print(f"{'Team':<22} {'16tel':>6} {'Achtel':>7} {'VF':>6} {'HF':>6} {'Finale':>7} {'WM-Sieg':>8}")
    for t in top:
        sc = team_stage_counts[t]
        grp = next(g for g, teams in groups.items() if t in teams)
        p_r32 = (group_pos_counts[grp][t][0] + group_pos_counts[grp][t][1]) / n + group_pos_counts[grp][t][2] / n * 0.67
        print(f"{t:<22} {pct(p_r32):>6} {pct(sc['p_R16']/n):>7} {pct(sc['p_QF']/n):>6} "
              f"{pct(sc['p_SF']/n):>6} {pct(sc['p_F']/n):>7} {pct(sc['p_Winner']/n):>8}")


if __name__ == "__main__":
    main()
